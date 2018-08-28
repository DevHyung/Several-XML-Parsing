import xml.etree.ElementTree as ET
import os, shutil
from openpyxl import Workbook, load_workbook
def log(tag, text):
	# Info tag
	if(tag == 'i'):
		print("[INFO] " + text)
	# Error tag
	elif(tag == 'e'):
		print("[ERROR] " + text)
	# Success tag
	elif(tag == 's'):
		print("[SUCCESS] " + text)
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for depth1List in _DATA:
            sheet.append(depth1List)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        book.save(_FILENAME)
def mkdir(_foldername):
    if not os.path.isdir("{}/".format(_foldername)):
        os.mkdir(_foldername)
        log('s',"폴더가 존재하지 않아 생성 되었습니다.")

if __name__ == "__main__":
    # === CONFIG
    srcFolder = "데이터" #이름만
    dstFolder = "완료"
    HEADER = ['piid id','branch','generic.name','generic.name.e']
    FILENAME = '결과.xlsx'
    # === Global Var
    dataList = []

    # === Code here
    save_excel(FILENAME, None, HEADER)  # init
    mkdir(srcFolder)
    mkdir(dstFolder)

    fileList = os.listdir('{}/'.format(srcFolder))
    #print(fileList)
    for file in fileList:
        # init
        xmlp = ET.XMLParser(encoding="utf-8")
        f = ET.parse('./{}/{}'.format(srcFolder,file),parser=xmlp)
        root = f.getroot()
        ppidList = []
        name = ''
        nameE = ''
        # parsing
        for parent in root.getiterator():
            for child in parent:
                if child.tag == 'generic.name':
                    name = child.text
                if child.tag == 'generic.name.e':
                    nameE = child.text
                if child.tag == 'piid':
                    ppidList.append(child.get('id')+','+child.get('branch'))
        # 전처리
        if name == '':
            log('e',"No Name")

        # 저장
        if len(dataList) >= 10: # 중간 저장
            save_excel(FILENAME, dataList, None)  # init
            dataList.clear()
        for ppid in ppidList:
            dataList.append([ppid.split(',')[0],ppid.split(',')[1],name,nameE])
        log('s',"{} 파일 완료".format(file))
        shutil.move("./{}/{}".format(srcFolder,file), "./{}/".format(dstFolder))
    save_excel(FILENAME, dataList, None)  # init